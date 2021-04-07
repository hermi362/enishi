# Simplify what is visible on OneDrive when the schedule xlsx is opened.
# After some trial and error, this is done by specifying a smallish print area.
# 2020-09 Herminio Gonzalez

from openpyxl import load_workbook
from openpyxl.utils import datetime as xldate
import datetime as dt
import sys

if len(sys.argv) == 1:
  print("usage: " + sys.argv[0] + " input_book.xlsx")
  exit

TODAYPY = dt.date.today()
# UGLY HACK! I had to force-set the year to 2019, because the sheet is fixed to that! <facepalm>
TODAYXL = xldate.to_excel(dt.datetime(2019, TODAYPY.month, TODAYPY.day, 0, 0))

filename = sys.argv[1]
print("Opening workbook " + filename)
wb = load_workbook(filename=filename, read_only=False)
ws = wb.active

startrow = 1
endrow = 1
loopdone = False
print("Searching for today's cell...")
for row in ws.iter_rows(min_row=1, max_row=2000, min_col=3, max_col=3):
  for cell in row:
    if cell.value == TODAYXL:
      print(cell, cell.value)
      startrow, endrow = cell.row - 14*4, cell.row + 14*4 - 1
      loopdone = True
  if loopdone:
    break

# from openpyxl import Workbook
# wb2 = Workbook()
# ws2 = wb2.active
# ws2.title = ws.title


# row2 = 10
# for row in ws.iter_rows(min_row=startrow, max_row=endrow, min_col=3, max_col=60):
#   col2 = 1
#   for cell in row:
#     ws2.cell(row=row2, column=col2, value=cell.value)
#     col2 = col2 + 1
#   row2 = row2 + 1

# wb2.save(filename = "out.xlsx")

print("Setting print area.")
ws.print_area = 'C'+str(startrow)+':'+'AF'+str(endrow)

print("Overwriting file.")
wb.save(filename = filename)
wb.close()
print("Finished.")

